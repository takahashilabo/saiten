import csv
import glob
import os
import sys
import shutil
import pathlib
import openpyxl
from PIL import Image  # 外部ライブラリ
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment

INPUT_DIR = "input"
OUTPUT_DIR = "output"
TRIMMED_FILE = "trimData.csv"
EXCEL_FILE = "saiten.xlsx"
EXCEL_MAG = 0.2 #Excelで表示する画像の縮小率（スキャン画像と画面解像度で見やすく調整する）

def readCSV(fname):
    if os.path.isfile(fname) == True:
        with open(fname) as f:
            reader = csv.reader(f)
            data = [row for row in reader]
            data.pop(0) #先頭（tag）を削除する
            return data
    else:
        print(f"{TRIMMED_FILE}のデータが不正です。斬る場所を決めてください。")
        sys.exit()

os.makedirs(INPUT_DIR, exist_ok=True)
shutil.rmtree(OUTPUT_DIR, ignore_errors=True) #出力ディレクトリ内容消去
#os.makedirs(OUTPUT_DIR, exist_ok=True)
#sys.exit()

data = readCSV(TRIMMED_FILE) #trimData.csv読み込み

files = [] #INPUT_DIR内の画像ファイル名取得
for name in sorted(os.listdir(INPUT_DIR)):
    if name.split(".")[-1] in ['jpg', "jpeg", "png", "PNG", "JPEG", "JPG", "gif"]:
        files.append(name)

wb = openpyxl.Workbook() #出力用Excel作成
ws = wb.worksheets[0]

row = 1 #Excelのセル番地
col = 1

for f in files:
    im = Image.open(os.path.join(INPUT_DIR, f)) #解答用紙画像

    print(f + "を斬ります" )

    for pos in data: #trimData.csvをもとに解答欄の画像を斬る
        title , left , top , right , bottom = pos
        
        outputDir = os.path.join(OUTPUT_DIR, title) #出力ディレクトリ
        os.makedirs(outputDir, exist_ok=True)

        #解答用紙画像から解答欄を切り抜きExcelでいい感じみ見えるようにリサイズ
        im_crop = im.crop((int(left), int(top), int(right), int(bottom)))
        im_crop_resize = im_crop.resize((int(im_crop.width * EXCEL_MAG), int(im_crop.height * EXCEL_MAG))) #Excelで表示したときの倍率（EXCEL_MAG）で調整
        
        im_crop_resize.save(os.path.join(outputDir, f), quality=95)

        #各設問の得点欄を作成＆調整
        o = ws.cell(row=row, column=col+1)
        o.value = 1 #１点という意味
        o.alignment = Alignment(horizontal='left', vertical='top')
        o.font = openpyxl.styles.fonts.Font(size=20)
        ws.column_dimensions[get_column_letter(col+1)].width = 5

        #解答欄画像に合わせて列幅調整＆セルに画像挿入
        cell_w, cell_h = im_crop_resize.size
        o = ws.column_dimensions[get_column_letter(col)]
        if (o.width < cell_w * 0.14):
            o.width = cell_w * 0.14
        img_to_excel = openpyxl.drawing.image.Image(os.path.join(outputDir, f))
        img_to_excel.anchor = ws.cell(row=row, column=col).coordinate
        ws.add_image(img_to_excel)
        col += 2
    ws.row_dimensions[row].height = cell_h 
    col = 1
    row += 1

print('全員分の解答用紙を斬りました。')

wb.save(EXCEL_FILE) #結果出力
